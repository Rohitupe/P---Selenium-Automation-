# write data in --> ExcelFiles/ReadData_16.1.xlsx

import openpyxl
excelPath = r'C:\Users\rohit\PycharmProjects\Selenium_Automation\Practice\DataDrivenTesting-Excel_16\ExcelFiles\WriteData_16.2.xlsx'

WorkBook = openpyxl.load_workbook(excelPath)

# print(WorkBook.sheetnames)
# Use sheet
sheet = WorkBook['Sheet1']

for r in range(1,6):
    for c in range(1,4):
        sheet.cell(row=r,column=c).value = "Rohit" # write data in excel file

WorkBook.save(excelPath) # save workbook once done writing

# check whether data is added or not
row = sheet.max_row
column = sheet.max_column

print(row,column)
