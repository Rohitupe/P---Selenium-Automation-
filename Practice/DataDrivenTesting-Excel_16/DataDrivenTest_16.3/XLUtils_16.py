import openpyxl

# get maximum number of Rows in the files
def getRowCount(file_path,sheet_name):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook[sheet_name]
    return sheet.max_row

# get maximum number of Columns in the files
def getColumnCount(file_path,sheet_name):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook[sheet_name]
    return sheet.max_column

# Read data from excel file
def readData(file_path,sheet_name,row_number,column_number):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook[sheet_name]
    return sheet.cell(row=row_number,column=column_number).value

# write data into excel file
def writeData(file_path,sheet_name,row_number,column_number,data):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook[sheet_name]
    sheet.cell(row=row_number,column=column_number).value = data
    workbook.save(file_path)
    return "Data Stored Successfully"