# read data from --> ExcelFiles/ReadData_16.1.xlsx

import openpyxl
excelPath = r'C:\Users\rohit\PycharmProjects\Selenium_Automation\Practice\DataDrivenTesting-Excel_16\ExcelFiles\ReadData_16.1.xlsx'

WorkbookFile = openpyxl.load_workbook(excelPath) # open excel file and store it in variable

# get sheets name present in excel file
# sheet = WorkbookFile.active() --> To get active sheet
# print(WorkbookFile.get_sheet_names())
print(WorkbookFile.sheetnames)

# select the sheet you want to work with
# FriendsSheet = WorkbookFile.get_sheet_by_name('Friends')
FriendsSheet = WorkbookFile['Friends']
# FriendsSheet = WorkbookFile.active

# read data by rows
rows = FriendsSheet.max_row
columns = FriendsSheet.max_column

for row in range(1,rows+1):
    for col in range(1,columns+1):
        print(FriendsSheet.cell(row=row,column=col).value,end="\t\t")
    print()
print(rows,columns)
